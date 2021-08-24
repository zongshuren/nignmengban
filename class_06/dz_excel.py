import time

from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self, butten=None):
        wb = load_workbook(self.file_name, read_only=True)
        sheet = wb[self.sheet_name]
        excel_list = []

        for i in range(2, sheet.max_row+1):
            excel_dict = {}
            try:
                excel_dict['ID'] = sheet.cell(i, 1).value
                excel_dict['data'] = sheet.cell(i, 2).value
                excel_dict['url'] = sheet.cell(i, 3).value
                excel_list.append(excel_dict)
            except TypeError as e:
                print('第{0}行出现错误请检查数据，错误为{1}'.format(i, e))
        if butten == None:
            final_data = excel_list
        else:
            final_data = []
            for item in excel_list:
                if item['ID'] in butten:
                    final_data.append(item)

        return final_data

if __name__ == '__main__':

    file = DoExcel(r'C:\Users\Administrator\Desktop\测试.xlsx', 'Sheet1').get_data()
    print(file)
#

# # 打开Excel文件
# wb = load_workbook(r'C:\Users\Administrator\Desktop\测试.xlsx')
# # 定位对应表单
# sheet = wb['Sheet1']
# # 最大行
# print('最大行：{0}'.format(sheet.max_row))
# # 最大列
# print('最大列：{0}'.format(sheet.max_column))
# hang = sheet.max_row
# lie = sheet.max_column
# # print(type(hang))
# for i in range(hang):
#     for j in range(lie):
#         # print(i+1, j+1)
#         res = sheet.cell(i+1, j+1).value
#         if res != None:
#             print(i+1, j+1, end=' ')
#             print(res)
#
# # 定位单元格，取出行列值
# res = sheet.cell(1, 1).value
#
# print(res)
