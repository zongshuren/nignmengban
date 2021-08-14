# -*- coding: utf-8 -*-
"""
@Author:ZongShuRen
@Time:2021/7/31 15:31
"""
import unittest
from HTMLTestRunner_PY3 import HTMLTestRunner
from class_06 import dz_case
import time
from class_06.dz_case import DzCase
from class_06.dz_excel import DoExcel
from openpyxl import load_workbook
import json

suit = unittest.TestSuite()
loder = unittest.TestLoader()

nowTime = time.strftime("%Y-%m-%d %H:%M", time.localtime())

# test_data = [
#     {'data': {"title": "自动化日常测试", "content": "自动化任务测试", "time": "2021-08-07 21:27", "end_time": "2021-08-07 21:27","employee": "[{\"label\":\"宗树仁\",\"value\":2127}]", "rolePlay": "consultant", "type_id": 4745, "status": 0},
#      'addCommonActivity_url': 'http://dztsl.lexue.com/api/activities/addCommonActivity'},
#
#     {'data': {"title": "自动化日常测试", "content": "自动化任务测试", "time": f"{nowTime}", "end_time": f"{nowTime}",
#                         "employee": "[{\"label\":\"宗树仁\",\"value\":2127}]", "rolePlay": "consultant", "type_id": 4745, "status": 0},
#      'addCommonActivity_url': 'http://dztsl.lexue.com/api/activities/addCommonActivity'}
# ]
#
test_data = DoExcel(r'C:\Users\Administrator\Desktop\测试.xlsx', 'Sheet1').get_data()

for item in test_data:
    print(item)
    try:
        suit.addTest(DzCase('test_dz_addCommonActivity', eval(item['data']), item['url']))
    except KeyError as e:
        print('数据出问题了，检查一下{0}'.format(e))

#
# for i in range(1, max_row+1):
#     try:
#         data = eval(sheet.cell(i, 1).value)
#         url = sheet.cell(i, 2).value
#         print(data, url)
#         suit.addTest(DzCase('test_dz_addCommonActivity', data, url))
#     except TypeError as e:
#         print('第{0}行出现错误请检查数据，错误为{1}'.format(i, e))

with open('dz_case.html', 'wb') as file:

    runner = HTMLTestRunner(stream=file, verbosity=2, title='多招测试用例', description='zsr')
    runner.run(suit)
