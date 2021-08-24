
from openpyxl import load_workbook


class DzExcel(object):

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        pass

    def get_header(self):
        # 打开Excel表格
        wb = load_workbook(self.file_name)
        # 找到工作簿
        sheet = wb[self.sheet_name]
        # 存储标题行
        header = []
        for j in range(1, sheet.max_column+1):
            header.append(sheet.cell(1, j).value)
        return header

    # def get_data(self):

        # wb = load_workbook(self.file_name)
        # sheet = wb[self.sheet_name]
        # header = self.get_header()
        # test_data = []
        # for i in range(2, sheet.max_row+1):
        #     sub_data = {}
        #     for j in range(1, sheet.max_column+1):
        #         print(j)
        #         sub_data[header[j-1]] = sheet.cell(i, j).value
        #     test_data.append(sub_data)
        # print(test_data)
        # return test_data

    def get_data2(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = self.get_header()
        list1 = []
        for i in range(2, sheet.max_row+1):
            zidian = {}
            for j in range(1, sheet.max_column+1):
                zidian[header[j-1]] = sheet.cell(i, j).value
            list1.append(zidian)
        return list1


if __name__ == '__main__':

    print(DzExcel(r'C:\Users\Administrator\Desktop\测试.xlsx', 'Sheet1').get_data2())
